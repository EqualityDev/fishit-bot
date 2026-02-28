import json
import random
import discord
from discord import app_commands
from discord.ext import commands
from datetime import datetime
from config import (
    STORE_NAME,
    STAFF_ROLE_NAME,
    DANA_NUMBER,
    BCA_NUMBER,

    CATEGORY_PRIORITY,
    STORE_THUMBNAIL,
    STORE_BANNER,
)
from utils import (
    calculate_total,
    format_items,
    send_invoice,
    save_products_json,
    load_products_json,
    is_staff,
)


class SpotlightModal(discord.ui.Modal, title="Buat Spotlight"):
    judul = discord.ui.TextInput(
        label="Judul",
        placeholder="Contoh: PROMO MINGGU INI!",
        max_length=100,
    )
    deskripsi = discord.ui.TextInput(
        label="Deskripsi",
        placeholder="Tulis kata-kata promosi di sini...",
        style=discord.TextStyle.paragraph,
        max_length=500,
    )
    gambar_url = discord.ui.TextInput(
        label="URL Gambar",
        placeholder="Contoh: https://i.imgur.com/xxx.png",
        max_length=500,
    )
    channel = discord.ui.TextInput(
        label="Channel (ketik ID atau nama channel)",
        placeholder="Contoh: 123456789 atau general",
        max_length=100,
    )

    async def on_submit(self, interaction: discord.Interaction):
        channel_input = self.channel.value.strip().replace("#", "").replace("<", "").replace(">", "")
        target_channel = None
        if channel_input.isdigit():
            target_channel = interaction.guild.get_channel(int(channel_input))
        else:
            target_channel = discord.utils.get(interaction.guild.text_channels, name=channel_input)

        if not target_channel:
            await interaction.response.send_message("❌ Channel tidak ditemukan!", ephemeral=True)
            return

        embed = discord.Embed(
            title=self.judul.value,
            description=self.deskripsi.value,
            color=0x00BFFF,
            timestamp=datetime.now(),
        )
        embed.set_image(url=self.gambar_url.value.strip())
        embed.set_thumbnail(url=STORE_THUMBNAIL)
        embed.set_footer(text=f"{STORE_NAME} • SPOTLIGHT", icon_url=STORE_THUMBNAIL)

        view = discord.ui.View()

        async def kirim_callback(btn_interaction: discord.Interaction):
            await target_channel.send(embed=embed)
            await btn_interaction.response.edit_message(
                content=f"✅ Spotlight berhasil dikirim ke {target_channel.mention}!",
                embed=None,
                view=None,
            )

        async def batal_callback(btn_interaction: discord.Interaction):
            await btn_interaction.response.edit_message(
                content="❌ Spotlight dibatalkan.",
                embed=None,
                view=None,
            )

        kirim_btn = discord.ui.Button(label="Kirim", style=discord.ButtonStyle.success)
        batal_btn = discord.ui.Button(label="Batal", style=discord.ButtonStyle.danger)
        kirim_btn.callback = kirim_callback
        batal_btn.callback = batal_callback
        view.add_item(kirim_btn)
        view.add_item(batal_btn)

        await interaction.response.send_message(
            content=f"**Preview spotlight** — cek dulu sebelum kirim ke {target_channel.mention}:",
            embed=embed,
            view=view,
            ephemeral=True,
        )


class StoreCog(commands.Cog):

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    # ─── Catalog ─────────────────────────────────────────────────

    @app_commands.command(name="catalog", description="Lihat semua item")
    async def catalog(self, interaction: discord.Interaction):
        await interaction.response.defer()
        products = await self.bot.products_cache.get_products()

        spotlight_items = [p for p in products if p.get("spotlight")]
        categories = {}
        for p in products:
            categories.setdefault(p["category"], []).append(p)

        all_cats = list(categories.keys())
        order = [c for c in CATEGORY_PRIORITY if c in all_cats]
        order += [c for c in all_cats if c not in order]

        embed = discord.Embed(
            title=f"{STORE_NAME} - READY STOCK",
            description="Payment: QRIS / DANA / BCA\n⚠️ Harga item Robux dapat berubah mengikuti rate pasar saat ini.",
            color=0x00BFFF,
        )
        embed.set_thumbnail(url=STORE_THUMBNAIL)
        embed.set_image(url=STORE_BANNER)

        if spotlight_items:
            value = "".join(
                f"**{p['name']}** — Rp {p['price']:,}\n"
                for p in spotlight_items
            )
            embed.add_field(name="SPOTLIGHT", value=value, inline=False)

        for cat in order:
            if cat in categories:
                items = categories[cat]
                value = "".join(
                    f"ID:{item['id']} - {item['name']} - Rp {item['price']:,}\n"
                    for item in items
                )
                embed.add_field(name=cat, value=value[:1024] or "-", inline=False)

        view = discord.ui.View()
        for cat in order:
            if cat in categories:
                view.add_item(discord.ui.Button(
                    label=f"BUY {cat}",
                    style=discord.ButtonStyle.primary,
                    custom_id=f"buy_{cat}",
                ))

        await interaction.followup.send(embed=embed, view=view)

    @app_commands.command(name="help", description=f"Bantuan menggunakan bot {STORE_NAME}")
    async def help_command(self, interaction: discord.Interaction):
        pages = [
            {
                "title": "CARA ORDER",
                "description": (
                    "**1.** `/catalog` → lihat semua item\n"
                    "**2.** Klik **BUY** → pilih item\n"
                    "**3.** Tiket terbuka otomatis\n"
                    "**4.** Ketik `1` / `2` / `3` pilih metode bayar\n"
                    "**5.** Transfer + kirim bukti pembayaran\n"
                    "**6.** Klik tombol **PAID**\n\n"
                    f"**Metode:** QRIS (1) · DANA {DANA_NUMBER} (2) · BCA {BCA_NUMBER} (3)"
                ),
            },
            {
                "title": "COMMAND CUSTOMER",
                "description": (
                    "`/catalog` — Lihat semua item\n"
                    "`/history` — Riwayat transaksi\n"
                    "`/items` — Item di tiket aktif\n"
                    "`/additem` — Tambah item ke tiket\n"
                    "`/removeitem` — Hapus item dari tiket\n"
                    "`/qris` — Lihat QR code\n"
                    "`!cancel` — Batalkan tiket"
                ),
            },
            {
                "title": "COMMAND ADMIN — PRODUK",
                "description": (
                    "`/addproduct` — Tambah produk\n"
                    "`/editprice` — Ubah harga\n"
                    "`/editname` — Ubah nama\n"
                    "`/deleteitem` — Hapus produk\n"
                    "`/importproduk` — Import dari Excel/CSV\n"
                    "`/uploadqris` — Upload QRIS\n"
                    "`/refreshcatalog` — Refresh catalog\n"
                    "`/spotlight` — Buat embed spotlight\n"
                    "`/setspotlight` — Set item spotlight (max 5)\n"
                    "`/unsetspotlight` — Hapus item spotlight\n"
                    "`/listspotlight` — Lihat item spotlight"
                ),
            },
            {
                "title": "COMMAND ADMIN — SISTEM",
                "description": (
                    "`/stats` — Statistik penjualan\n"
                    "`/statdetail` — Detail statistik\n"
                    "`/allhistory` — Semua transaksi user\n"
                    "`/transcript` — Cari transcript tiket\n"
                    "`/export` — Export data CSV\n"
                    "`/broadcast` — Kirim pesan ke semua\n"
                    "`/blacklist` — Blokir user\n"
                    "`/unblacklist` — Hapus blokir user\n"
                    "`/backup` — Backup manual DB\n"
                    "`/listbackup` — Daftar backup\n"
                    "`/restore` — Restore backup\n"
                    "`/migrate` — Export/import data migrasi\n"
                    "`/resetdb` — Reset database\n"
                    "`/cleanupstats` — Hapus data statistik\n"
                    "`/fakeinvoice` — Generate invoice test\n"
                    "`/ping` — Cek status bot\n"
                    "`/reboot` — Restart bot"
                ),
            },
            {
                "title": "COMMAND ADMIN — LAINNYA",
                "description": (
                    "**Giveaway**\n"
                    "`/giveaway` — Mulai giveaway baru\n"
                    "`/giveaway_end` — Akhiri giveaway lebih awal\n"
                    "`/giveaway_reroll` — Reroll pemenang\n"
                    "`/giveaway_list` — Lihat giveaway aktif\n\n"
                    "**Auto React**\n"
                    "`/setreact` — Set auto react (staff)\n"
                    "`/setreactall` — Set auto react (semua)\n"
                    "`/reactlist` — Lihat daftar react"
                ),
            },
        ]

        def build_embed(page_index):
            page = pages[page_index]
            embed = discord.Embed(
                title=page["title"],
                description=page["description"],
                color=0x00BFFF,
            )
            embed.set_thumbnail(url=STORE_THUMBNAIL)
            embed.set_footer(text=f"{STORE_NAME} • Halaman {page_index + 1}/{len(pages)}", icon_url=STORE_THUMBNAIL)
            return embed

        class HelpView(discord.ui.View):
            def __init__(self):
                super().__init__(timeout=60)
                self.page = 0

            @discord.ui.button(label="◀ Prev", style=discord.ButtonStyle.secondary)
            async def prev_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                self.page = (self.page - 1) % len(pages)
                await interaction.response.edit_message(embed=build_embed(self.page), view=self)

            @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.secondary)
            async def next_button(self, interaction: discord.Interaction, button: discord.ui.Button):
                self.page = (self.page + 1) % len(pages)
                await interaction.response.edit_message(embed=build_embed(self.page), view=self)

        await interaction.response.send_message(embed=build_embed(0), view=HelpView())

    # ─── Product Management ──────────────────────────────────────

    @app_commands.command(name="addproduct", description="[ADMIN] Tambah produk baru")
    @app_commands.describe(id="ID produk (angka unik)", name="Nama produk", price="Harga", category="Kategori")
    async def add_product(self, interaction: discord.Interaction, id: int, name: str, price: int, category: str):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        if any(p["id"] == id for p in self.bot.PRODUCTS):
            await interaction.response.send_message(f"❌ ID {id} sudah dipakai!", ephemeral=True)
            return
        if price <= 0:
            await interaction.response.send_message("❌ Harga harus lebih dari 0!", ephemeral=True)
            return
        new_product = {"id": id, "name": name, "price": price, "category": category.upper()}
        self.bot.PRODUCTS.append(new_product)
        save_products_json(self.bot.PRODUCTS)
        await self.bot.db.save_products(self.bot.PRODUCTS)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="✅ PRODUK DITAMBAHKAN",
            description=f"**ID:** {id}\n**Nama:** {name}\n**Harga:** Rp {price:,}\n**Kategori:** {category.upper()}",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="editprice", description="[ADMIN] Ubah harga item")
    @app_commands.describe(item_id="ID item", new_price="Harga baru")
    async def edit_price(self, interaction: discord.Interaction, item_id: int, new_price: int):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        if new_price <= 0:
            await interaction.response.send_message("❌ Harga harus lebih dari 0!", ephemeral=True)
            return
        old_price = item["price"]
        item["price"] = new_price
        save_products_json(self.bot.PRODUCTS)
        await self.bot.db.save_products(self.bot.PRODUCTS)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="💰 HARGA DIUPDATE",
            description=f"**Item:** {item['name']}\n**Lama:** Rp {old_price:,}\n**Baru:** Rp {new_price:,}",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="editname", description="[ADMIN] Ubah nama item")
    @app_commands.describe(item_id="ID item", new_name="Nama baru")
    async def edit_name(self, interaction: discord.Interaction, item_id: int, new_name: str):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        old_name = item["name"]
        item["name"] = new_name
        save_products_json(self.bot.PRODUCTS)
        await self.bot.db.save_products(self.bot.PRODUCTS)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="📝 NAMA DIUPDATE",
            description=f"**ID:** {item_id}\n**Lama:** {old_name}\n**Baru:** {new_name}",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="deleteitem", description="[ADMIN] Hapus item")
    @app_commands.describe(item_id="ID item yang akan dihapus")
    async def delete_item(self, interaction: discord.Interaction, item_id: int):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        self.bot.PRODUCTS.remove(item)
        save_products_json(self.bot.PRODUCTS)
        await self.bot.db.save_products(self.bot.PRODUCTS)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="🗑️ ITEM DIHAPUS",
            description=f"**ID:** {item_id}\n**Nama:** {item['name']}\n**Harga:** Rp {item['price']:,}",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="spotlight", description="[ADMIN] Buat dan kirim embed spotlight")
    async def spotlight(self, interaction: discord.Interaction):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        await interaction.response.send_modal(SpotlightModal())

    @app_commands.command(name="setspotlight", description="[ADMIN] Set produk sebagai spotlight")
    @app_commands.describe(item_id="ID item yang akan di-spotlight")
    async def set_spotlight(self, interaction: discord.Interaction, item_id: int):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        # Limit spotlight maksimal 5
        spotlight_count = sum(1 for p in self.bot.PRODUCTS if p.get("spotlight"))
        if spotlight_count >= 5 and not item.get("spotlight"):
            await interaction.response.send_message("❌ Maksimal 5 produk spotlight! Hapus salah satu dulu dengan /unsetspotlight.", ephemeral=True)
            return
        item["spotlight"] = 1
        await self.bot.db.set_spotlight(item_id, 1)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="SPOTLIGHT DIAKTIFKAN",
            description=f"**{item['name']}** sekarang tampil di spotlight catalog!",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="unsetspotlight", description="[ADMIN] Hapus produk dari spotlight")
    @app_commands.describe(item_id="ID item yang akan dihapus dari spotlight")
    async def unset_spotlight(self, interaction: discord.Interaction, item_id: int):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        item["spotlight"] = 0
        await self.bot.db.set_spotlight(item_id, 0)
        self.bot.products_cache.invalidate()
        embed = discord.Embed(
            title="🔦 SPOTLIGHT DINONAKTIFKAN",
            description=f"**{item['name']}** dihapus dari spotlight.",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="listspotlight", description="[ADMIN] Lihat produk yang sedang spotlight")
    async def list_spotlight(self, interaction: discord.Interaction):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        products = await self.bot.products_cache.get_products()
        spotlight = [p for p in products if p.get("spotlight")]
        if not spotlight:
            await interaction.response.send_message("📝 Belum ada produk spotlight.", ephemeral=True)
            return
        embed = discord.Embed(title="🔦 PRODUK SPOTLIGHT", color=0x00BFFF)
        for p in spotlight:
            embed.add_field(name=f"ID:{p['id']} — {p['name']}", value=f"Rp {p['price']:,}", inline=False)
        await interaction.response.send_message(embed=embed)
    async def refresh_cache(self, interaction: discord.Interaction):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        await self.bot.products_cache.refresh()
        await interaction.response.send_message(
            f"✅ Cache refreshed! {len(self.bot.products_cache.data)} products loaded"
        )

    @app_commands.command(name="importproduk", description="[ADMIN] Import produk dari file Excel/CSV")
    @app_commands.describe(file="Upload file .xlsx atau .csv")
    async def import_produk(self, interaction: discord.Interaction, file: discord.Attachment):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return

        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
            await interaction.response.send_message("❌ Format file harus `.xlsx` atau `.csv`!", ephemeral=True)
            return

        await interaction.response.defer()

        try:
            import io
            file_bytes = await file.read()

            products = []
            errors = []

            if filename.endswith(".csv"):
                import csv
                text = file_bytes.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                for i, row in enumerate(reader, start=2):
                    try:
                        products.append({
                            "id": int(str(row["id"]).strip()),
                            "name": str(row["name"]).strip(),
                            "price": int(str(row["price"]).strip().replace(",", "").replace(".", "")),
                            "category": str(row["category"]).strip().upper(),
                            "spotlight": 0,
                        })
                    except Exception as e:
                        errors.append(f"Baris {i}: {e}")

            elif filename.endswith(".xlsx"):
                try:
                    import openpyxl
                except ImportError:
                    await interaction.followup.send("❌ Library `openpyxl` belum terinstall di server. Jalankan: `pip install openpyxl --break-system-packages` di Termux.")
                    return

                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    data = dict(zip(headers, row))
                    if not data.get("id"):
                        continue
                    try:
                        products.append({
                            "id": int(data["id"]),
                            "name": str(data["name"]).strip(),
                            "price": int(str(data["price"]).replace(",", "").replace(".", "")),
                            "category": str(data["category"]).strip().upper(),
                            "spotlight": 0,
                        })
                    except Exception as e:
                        errors.append(f"Baris {i}: {e}")

            if not products:
                await interaction.followup.send("❌ Tidak ada produk yang bisa diimport. Pastikan format kolom: `id`, `name`, `price`, `category`")
                return

            # Merge ke PRODUCTS yang ada
            added, updated = 0, 0
            existing_ids = {p["id"]: i for i, p in enumerate(self.bot.PRODUCTS)}
            for p in products:
                if p["id"] in existing_ids:
                    self.bot.PRODUCTS[existing_ids[p["id"]]].update({
                        "name": p["name"], "price": p["price"], "category": p["category"]
                    })
                    updated += 1
                else:
                    self.bot.PRODUCTS.append(p)
                    added += 1

            self.bot.PRODUCTS.sort(key=lambda x: x["id"])
            save_products_json(self.bot.PRODUCTS)
            await self.bot.db.save_products(self.bot.PRODUCTS)
            self.bot.products_cache.invalidate()

            desc = f"✅ **{added}** produk ditambahkan\n✏️ **{updated}** produk diupdate\n📦 Total: **{len(self.bot.PRODUCTS)}** produk"
            if errors:
                desc += f"\n\n⚠️ **{len(errors)} baris dilewati:**\n" + "\n".join(errors[:5])

            embed = discord.Embed(title="IMPORT PRODUK SELESAI", description=desc, color=0x00BFFF)
            await interaction.followup.send(embed=embed)

        except Exception as e:
            await interaction.followup.send(f"❌ Error: {e}")

    @app_commands.command(name="refreshcatalog", description="[ADMIN] Refresh catalog tanpa restart")
    async def refresh_catalog(self, interaction: discord.Interaction):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        self.bot.PRODUCTS = load_products_json()
        embed = discord.Embed(
            title="🔄 CATALOG REFRESHED",
            description=f"Total item: {len(self.bot.PRODUCTS)}",
            color=0x00BFFF,
        )
        await interaction.response.send_message(embed=embed)

    # ─── QRIS ────────────────────────────────────────────────────

    @app_commands.command(name="uploadqris", description="[ADMIN] Upload QRIS")
    @app_commands.describe(image="Upload file gambar QR code")
    async def upload_qris(self, interaction: discord.Interaction, image: discord.Attachment):
        if not is_staff(interaction):
            await interaction.response.send_message("Admin only!", ephemeral=True)
            return
        if not image.content_type.startswith("image/"):
            await interaction.response.send_message("File harus gambar!", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True)
        qr_channel = discord.utils.get(interaction.guild.channels, name="qr-code")
        if not qr_channel:
            qr_channel = await interaction.guild.create_text_channel(
                name="qr-code",
                overwrites={
                    interaction.guild.default_role: discord.PermissionOverwrite(read_messages=True, send_messages=False),
                    interaction.guild.me: discord.PermissionOverwrite(read_messages=True, send_messages=True),
                },
            )
        embed = discord.Embed(title="QRIS PAYMENT", color=0x00BFFF)
        embed.set_image(url=image.url)
        embed.set_footer(text=f"Uploaded by {interaction.user.name}")
        await qr_channel.send(embed=embed)
        await interaction.followup.send(f"QRIS uploaded to {qr_channel.mention}", ephemeral=True)

    @app_commands.command(name="qris", description="Lihat QR code")
    async def cek_qris(self, interaction: discord.Interaction):
        await interaction.response.defer()
        qr_channel = discord.utils.get(interaction.guild.channels, name="qr-code")
        if not qr_channel:
            await interaction.followup.send("QR code tidak tersedia!", ephemeral=True)
            return
        async for msg in qr_channel.history(limit=10):
            if msg.author == self.bot.user and msg.embeds:
                await interaction.followup.send(embed=msg.embeds[0])
                return
        await interaction.followup.send("QR code tidak ditemukan!", ephemeral=True)

    # ─── History ─────────────────────────────────────────────────

    @app_commands.command(name="history", description="Lihat riwayat transaksi pribadi")
    async def history(self, interaction: discord.Interaction):
        user_id = str(interaction.user.id)
        last_5 = await self.bot.db.get_user_transactions(user_id, limit=5)
        if not last_5:
            await interaction.response.send_message("Belum ada transaksi.", ephemeral=True)
            return
        all_user = await self.bot.db.get_user_transactions(user_id, limit=1000)
        embed = discord.Embed(
            title="RIWAYAT TRANSAKSI",
            description=f"Total: {len(all_user)} transaksi",
            color=0x00BFFF,
        )
        for t in reversed(last_5):
            ts = t["timestamp"] if isinstance(t["timestamp"], datetime) else datetime.fromisoformat(t["timestamp"])
            items = t["items"] if isinstance(t["items"], list) else json.loads(t["items"])
            items_short = ", ".join(f"{i['qty']}x {i['name'][:15]}" for i in items[:2])
            if len(items) > 2:
                items_short += f" +{len(items)-2} lagi"
            embed.add_field(
                name=f"{t['invoice']} - {ts.strftime('%d/%m/%Y %H:%M')}",
                value=f"{items_short} | Rp {t['total_price']:,} | {t.get('payment_method', '-')}",
                inline=False,
            )
        await interaction.response.send_message(embed=embed, ephemeral=True)

    @app_commands.command(name="allhistory", description="[ADMIN] Lihat SEMUA riwayat transaksi user")
    @app_commands.describe(user="User yang mau dilihat")
    async def all_history(self, interaction: discord.Interaction, user: discord.User):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        all_trans = await self.bot.db.get_user_transactions(str(user.id), limit=1000)
        if not all_trans:
            await interaction.response.send_message(f"📝 {user.mention} belum punya transaksi.", ephemeral=True)
            return
        total_spent = sum(t["total_price"] for t in all_trans)
        embed = discord.Embed(
            title=f"📋 SEMUA TRANSAKSI {user.name}",
            description=f"Total: **{len(all_trans)}** | Belanja: **Rp {total_spent:,}**",
            color=0x00BFFF,
        )
        for t in all_trans[-10:]:
            ts = t["timestamp"] if isinstance(t["timestamp"], datetime) else datetime.fromisoformat(t["timestamp"])
            items = t["items"] if isinstance(t["items"], list) else json.loads(t["items"])
            items_short = ", ".join(f"{i['qty']}x {i['name'][:15]}" for i in items[:2])
            if len(items) > 2:
                items_short += f" +{len(items)-2} lagi"
            embed.add_field(
                name=f"{t['invoice']} - {ts.strftime('%d/%m/%Y %H:%M')}",
                value=f"{items_short} | Rp {t['total_price']:,} | {t.get('payment_method', '-')}",
                inline=False,
            )
        if len(all_trans) > 10:
            embed.set_footer(text=f"Menampilkan 10 dari {len(all_trans)} transaksi")
        await interaction.response.send_message(embed=embed)

    # ─── Fake Invoice ────────────────────────────────────────────

    @app_commands.command(name="fakeinvoice", description="[ADMIN] Generate fake invoice")
    @app_commands.describe(jumlah="Jumlah invoice (1-5)")
    async def fake_invoice(self, interaction: discord.Interaction, jumlah: int = 1):
        if not is_staff(interaction):
            await interaction.response.send_message("❌ Admin only!", ephemeral=True)
            return
        if not (1 <= jumlah <= 5):
            await interaction.response.send_message("❌ Jumlah minimal 1, maksimal 5", ephemeral=True)
            return
        await interaction.response.send_message(f"🧪 Generating {jumlah} fake invoice...", ephemeral=True, delete_after=3)
        methods = ["DANA", "BCA", "QRIS"]
        weights = [0.5, 0.3, 0.2]
        for _ in range(jumlah):
            num_items = random.choices([1, 2, 3], weights=[0.6, 0.3, 0.1])[0]
            selected = random.sample(self.bot.PRODUCTS, k=min(num_items, len(self.bot.PRODUCTS)))
            items = [
                {"id": p["id"], "name": p["name"], "price": p["price"], "qty": random.randint(1, 3)}
                for p in selected
            ]
            await send_invoice(
                interaction.guild,
                {
                    "user_id": str(random.randint(100000000000000000, 999999999999999999)),
                    "items": items,
                    "total_price": sum(i["price"] * i["qty"] for i in items),
                    "payment_method": random.choices(methods, weights=weights)[0],
                    "admin_id": str(interaction.user.id),
                    "fake": True,
                },
                self.bot.db,
            )
        await interaction.followup.send(f"✅ {jumlah} fake invoice dikirim ke channel log!", ephemeral=True)

    # ─── Ticket Item Commands ────────────────────────────────────

    @app_commands.command(name="additem", description="➕ Tambah item ke tiket ini")
    @app_commands.describe(item_id="ID item", qty="Jumlah (default 1)")
    async def add_item_to_ticket(self, interaction: discord.Interaction, item_id: int, qty: int = 1):
        if not interaction.channel.name.startswith("ticket-"):
            await interaction.response.send_message("❌ Ini bukan channel tiket!", ephemeral=True)
            return
        channel_id = str(interaction.channel.id)
        if channel_id not in self.bot.active_tickets or self.bot.active_tickets[channel_id]["status"] != "OPEN":
            await interaction.response.send_message("❌ Tiket tidak ditemukan atau sudah closed!", ephemeral=True)
            return
        item = next((p for p in self.bot.PRODUCTS if p["id"] == item_id), None)
        if not item:
            await interaction.response.send_message("❌ Item tidak ditemukan!", ephemeral=True)
            return
        ticket = self.bot.active_tickets[channel_id]
        staff_role = discord.utils.get(interaction.guild.roles, name=STAFF_ROLE_NAME)
        if str(interaction.user.id) != ticket["user_id"] and staff_role not in interaction.user.roles:
            await interaction.response.send_message("❌ Hanya pemilik tiket atau admin!", ephemeral=True)
            return
        found = False
        for existing in ticket["items"]:
            if existing["id"] == item_id:
                existing["qty"] += qty
                found = True
                break
        if not found:
            ticket["items"].append({"id": item["id"], "name": item["name"], "price": item["price"], "qty": qty})
        ticket["total_price"] = calculate_total(ticket["items"])
        await self.bot.db.update_ticket_items(channel_id, ticket["items"])
        await self.bot.db.update_ticket_total(channel_id, ticket["total_price"])
        embed = discord.Embed(
            title="➕ ITEM DITAMBAHKAN",
            description=f"**{qty}x {item['name']}** berhasil ditambahkan!",
            color=0x00BFFF,
        )
        embed.add_field(name="🛒 ITEMS SAAT INI", value=format_items(ticket["items"]), inline=False)
        embed.add_field(name="💰 TOTAL", value=f"Rp {ticket['total_price']:,}", inline=False)
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="removeitem", description="➖ Hapus item dari tiket ini")
    @app_commands.describe(item_id="ID item", qty="Jumlah yang dihapus (default semua)")
    async def remove_item_from_ticket(self, interaction: discord.Interaction, item_id: int, qty: int = None):
        if not interaction.channel.name.startswith("ticket-"):
            await interaction.response.send_message("❌ Ini bukan channel tiket!", ephemeral=True)
            return
        channel_id = str(interaction.channel.id)
        if channel_id not in self.bot.active_tickets or self.bot.active_tickets[channel_id]["status"] != "OPEN":
            await interaction.response.send_message("❌ Tiket tidak ditemukan!", ephemeral=True)
            return
        ticket = self.bot.active_tickets[channel_id]
        staff_role = discord.utils.get(interaction.guild.roles, name=STAFF_ROLE_NAME)
        if str(interaction.user.id) != ticket["user_id"] and staff_role not in interaction.user.roles:
            await interaction.response.send_message("❌ Hanya pemilik tiket atau admin!", ephemeral=True)
            return
        item_found = next((item for item in ticket["items"] if item["id"] == item_id), None)
        if not item_found:
            await interaction.response.send_message("❌ Item tidak ditemukan di tiket!", ephemeral=True)
            return
        if qty is None or qty >= item_found["qty"]:
            ticket["items"].remove(item_found)
            removal_msg = f"✅ **{item_found['qty']}x {item_found['name']}** dihapus dari tiket!"
        else:
            item_found["qty"] -= qty
            removal_msg = f"✅ **{qty}x {item_found['name']}** dikurangi!\nSisa: {item_found['qty']}x"
        ticket["total_price"] = calculate_total(ticket["items"])
        await self.bot.db.update_ticket_items(channel_id, ticket["items"])
        await self.bot.db.update_ticket_total(channel_id, ticket["total_price"])
        if not ticket["items"]:
            await interaction.response.send_message("🔄 Tiket kosong, menutup tiket dalam 5 detik...")
            import asyncio
            await asyncio.sleep(5)
            del self.bot.active_tickets[channel_id]
            await self.bot.db.delete_ticket(channel_id)
            await interaction.channel.delete()
            return
        embed = discord.Embed(title="➖ ITEM DIHAPUS", description=removal_msg, color=0x00BFFF)
        embed.add_field(name="🛒 ITEMS SAAT INI", value=format_items(ticket["items"]), inline=False)
        embed.add_field(name="💰 TOTAL", value=f"Rp {ticket['total_price']:,}", inline=False)
        await interaction.response.send_message(embed=embed)

    @app_commands.command(name="items", description="Lihat item di tiket ini")
    async def list_items(self, interaction: discord.Interaction):
        if not interaction.channel.name.startswith("ticket-"):
            await interaction.response.send_message("❌ Ini bukan channel tiket!", ephemeral=True)
            return
        channel_id = str(interaction.channel.id)
        if channel_id not in self.bot.active_tickets:
            await interaction.response.send_message("❌ Tiket tidak ditemukan!", ephemeral=True)
            return
        ticket = self.bot.active_tickets[channel_id]
        embed = discord.Embed(
            title="🛒 DAFTAR ITEM",
            description=format_items(ticket["items"]) or "Belum ada item",
            color=0x00BFFF,
        )
        embed.add_field(name="💰 TOTAL", value=f"Rp {ticket['total_price']:,}", inline=False)
        await interaction.response.send_message(embed=embed, ephemeral=True)


async def setup(bot: commands.Bot):
    await bot.add_cog(StoreCog(bot))
