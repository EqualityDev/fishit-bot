"""
Cellyn Store — Product Importer
================================
Import produk dari file CSV atau Excel ke database bot.

Cara pakai:
  python3 import_products.py products.csv
  python3 import_products.py products.xlsx

Format kolom yang diperlukan:
  id | name | price | category

Contoh isi file:
  1, Nitro 1 Month, 75000, NITRO
  2, Robux 1000, 85000, ROBUX
"""

import sys
import json
import sqlite3
from pathlib import Path

DB_NAME = "store.db"
PRODUCTS_JSON = "products.json"


def load_from_csv(filepath):
    import csv
    products = []
    with open(filepath, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                products.append({
                    "id": int(row["id"].strip()),
                    "name": row["name"].strip(),
                    "price": int(str(row["price"]).strip().replace(",", "").replace(".", "")),
                    "category": row["category"].strip().upper(),
                    "spotlight": 0,
                })
            except Exception as e:
                print(f"⚠️  Skip baris {row}: {e}")
    return products


def load_from_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl tidak terinstall. Jalankan: pip install openpyxl --break-system-packages")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [str(cell.value).strip().lower() for cell in ws[1]]

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get("id"):
            continue
        try:
            products.append({
                "id": int(data["id"]),
                "name": str(data["name"]).strip(),
                "price": int(str(data["price"]).replace(",", "").replace(".", "")),
                "category": str(data["category"]).strip().upper(),
                "spotlight": 0,
            })
        except Exception as e:
            print(f"⚠️  Skip baris {data}: {e}")
    return products


def import_to_db(products):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    added = 0
    updated = 0
    for p in products:
        cur.execute("SELECT id FROM products WHERE id = ?", (p["id"],))
        exists = cur.fetchone()
        if exists:
            cur.execute(
                "UPDATE products SET name=?, price=?, category=? WHERE id=?",
                (p["name"], p["price"], p["category"], p["id"])
            )
            updated += 1
        else:
            cur.execute(
                "INSERT INTO products (id, name, price, category, spotlight) VALUES (?, ?, ?, ?, ?)",
                (p["id"], p["name"], p["price"], p["category"], p["spotlight"])
            )
            added += 1

    conn.commit()
    conn.close()
    return added, updated


def save_json(products):
    # Merge dengan JSON yang ada
    existing = []
    if Path(PRODUCTS_JSON).exists():
        with open(PRODUCTS_JSON) as f:
            existing = json.load(f)

    existing_ids = {p["id"]: i for i, p in enumerate(existing)}
    for p in products:
        if p["id"] in existing_ids:
            existing[existing_ids[p["id"]]].update(p)
        else:
            existing.append(p)

    existing.sort(key=lambda x: x["id"])
    with open(PRODUCTS_JSON, "w") as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)


def main():
    if len(sys.argv) < 2:
        print("Cara pakai: python3 import_products.py <file.csv atau file.xlsx>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"❌ File tidak ditemukan: {filepath}")
        sys.exit(1)

    ext = Path(filepath).suffix.lower()
    print(f"📂 Membaca file: {filepath}")

    if ext == ".csv":
        products = load_from_csv(filepath)
    elif ext in [".xlsx", ".xls"]:
        products = load_from_excel(filepath)
    else:
        print("❌ Format file tidak didukung. Gunakan .csv atau .xlsx")
        sys.exit(1)

    if not products:
        print("❌ Tidak ada produk yang bisa diimport.")
        sys.exit(1)

    print(f"✅ {len(products)} produk berhasil dibaca.")
    print("")

    # Preview
    print(f"{'ID':<5} {'Nama':<30} {'Harga':<12} {'Kategori'}")
    print("-" * 60)
    for p in products[:10]:
        print(f"{p['id']:<5} {p['name']:<30} Rp {p['price']:,<10} {p['category']}")
    if len(products) > 10:
        print(f"... dan {len(products) - 10} produk lainnya")

    print("")
    confirm = input("Import ke database? (y/n): ").strip().lower()
    if confirm != "y":
        print("❌ Import dibatalkan.")
        sys.exit(0)

    added, updated = import_to_db(products)
    save_json(products)

    print("")
    print(f"✅ Import selesai!")
    print(f"   ➕ Ditambahkan : {added} produk baru")
    print(f"   ✏️  Diupdate    : {updated} produk")
    print("")
    print("Restart bot untuk menerapkan perubahan: python3 bot.py")


if __name__ == "__main__":
    main()
